"""
tc-table handler — Table data processor for path pipelines.

Read CSV/TSV/XLSX → JSON arrays → filter/sort/pivot/join → write back.
Zero external deps for CSV/TSV. openpyxl optional for XLSX.

Directives:
    tc-table;read,<path>[,<format>]         — read file → JSON
    tc-table;schema,<path>                  — columns + types + samples
    tc-table;filter,<data>,<JSON>           — filter rows
    tc-table;sort,<data>,<JSON>             — sort rows
    tc-table;pivot,<data>,<JSON>            — group + aggregate
    tc-table;join,<data_a>,<data_b>,<JSON>  — join two tables
    tc-table;write,<data>,<path>[,<format>] — write to file
"""
import csv
import io
import json
import logging
import operator
from pathlib import Path

from core.registry import directive

logger = logging.getLogger(__name__)

_project_root: Path | None = None


def init_tc_table_handler(project_root: str = None):
    global _project_root
    if project_root:
        _project_root = Path(project_root)
    logger.info("tc-table initialised")


def _resolve_path(file_path: str) -> Path:
    p = Path(file_path)
    if p.is_absolute():
        return p
    if _project_root:
        return _project_root / p
    return p.resolve()


def _detect_format(file_path: str, fmt: str = None) -> str:
    if fmt:
        return fmt.lower()
    path = file_path.lower()
    if path.endswith(".tsv"):
        return "tsv"
    if path.endswith(".xlsx"):
        return "xlsx"
    return "csv"


def _parse_json_params(params: list[str], start_idx: int = 1) -> dict:
    if len(params) <= start_idx:
        raise ValueError("missing JSON parameter")
    direct = params[start_idx]
    try:
        return json.loads(direct)
    except json.JSONDecodeError:
        pass
    joined = ",".join(params[start_idx:])
    return json.loads(joined)


def _parse_json_rows(json_str: str) -> list[dict]:
    try:
        data = json.loads(json_str)
    except json.JSONDecodeError:
        return []
    if isinstance(data, dict) and "rows" in data:
        return data["rows"]
    if isinstance(data, list):
        return data
    return []



def _read_csv(path: str, delimiter: str) -> list[dict]:
    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        return [dict(row) for row in reader]


def _read_xlsx(path: str) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl required for .xlsx files. pip install openpyxl")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(cell.value) if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append({headers[i]: str(v) if v is not None else "" for i, v in enumerate(row)})
    wb.close()
    return rows


def _read_file(file_path: str, fmt: str = None) -> tuple[list[dict], str]:
    fmt = _detect_format(file_path, fmt)
    path = _resolve_path(file_path)
    if fmt == "tsv":
        rows = _read_csv(str(path), "\t")
    elif fmt == "xlsx":
        rows = _read_xlsx(str(path))
    else:
        rows = _read_csv(str(path), ",")
    return rows, fmt



@directive("tc-table", "read", domain_alias="表格", action_aliases={"read": "读取"})
def tc_table_read(params: list[str]) -> dict:
    if not params:
        return {"status": "error", "reason": "Usage: tc-table;read,<file_path>[,<format>]"}

    file_path = params[0]
    fmt = params[1] if len(params) > 1 else None

    try:
        rows, detected = _read_file(file_path, fmt)
        columns = list(rows[0].keys()) if rows else []
        return {
            "status": "ok",
            "source": file_path,
            "format": detected,
            "columns": columns,
            "rows": rows,
            "count": len(rows),
        }
    except FileNotFoundError:
        return {"status": "error", "reason": f"file not found: {file_path}"}
    except Exception as e:
        logger.exception("tc-table;read failed: %s", file_path)
        return {"status": "error", "reason": str(e)}



def _infer_type(values: list[str]) -> str:
    if not values:
        return "empty"
    all_num = True
    for v in values:
        if not v or v == "":
            continue
        try:
            float(v)
        except ValueError:
            all_num = False
            break
    return "number" if all_num else "string"


@directive("tc-table", "schema", domain_alias="表格", action_aliases={"schema": "结构"})
def tc_table_schema(params: list[str]) -> dict:
    if not params:
        return {"status": "error", "reason": "Usage: tc-table;schema,<file_path>"}

    file_path = params[0]
    try:
        rows, _ = _read_file(file_path)
        if not rows:
            return {"status": "ok", "source": file_path, "columns": [], "row_count": 0}

        columns = []
        sample_rows = rows[:100]
        for col in rows[0].keys():
            col_values = [row.get(col, "") for row in sample_rows]
            samples = [v for v in col_values[:3] if v and v != ""]
            columns.append({
                "name": col,
                "inferred_type": _infer_type(col_values),
                "sample_values": samples,
            })

        return {
            "status": "ok",
            "source": file_path,
            "columns": columns,
            "row_count": len(rows),
        }
    except FileNotFoundError:
        return {"status": "error", "reason": f"file not found: {file_path}"}
    except Exception as e:
        logger.exception("tc-table;schema failed: %s", file_path)
        return {"status": "error", "reason": str(e)}



_OPS = {
    "=": operator.eq,
    "!=": operator.ne,
    ">": operator.gt,
    "<": operator.lt,
    ">=": operator.ge,
    "<=": operator.le,
}


def _compare(a: str, b: str, op: str) -> bool:
    if op in _OPS:
        try:
            return _OPS[op](float(a), float(b))
        except (ValueError, TypeError):
            return _OPS[op](a, b)
    if op == "contains":
        return b in a
    if op == "starts":
        return a.startswith(b)
    if op == "ends":
        return a.endswith(b)
    if op == "in":
        return a in set(x.strip() for x in b.split(","))
    raise ValueError(f"unknown operator: {op}")



@directive("tc-table", "filter", domain_alias="表格", action_aliases={"filter": "筛选"})
def tc_table_filter(params: list[str]) -> dict:
    if len(params) < 2:
        return {"status": "error", "reason": "Usage: tc-table;filter,<json_data>,<JSON_query>"}

    rows = _parse_json_rows(params[0])
    if not rows:
        return {"status": "error", "reason": "no data rows provided or JSON parse failed"}
    try:
        query = json.loads(params[1])
    except json.JSONDecodeError:
        return {"status": "error", "reason": f"invalid JSON query: {params[1][:50]}"}

    where = query.get("where")
    if not where or not isinstance(where, list) or len(where) != 3:
        return {"status": "error", "reason": "query must have 'where': [col, op, val]"}

    col, op, val = where
    limit = min(int(query.get("limit", 0)), 1000) or 0
    total = len(rows)

    try:
        filtered = [r for r in rows if _compare(r.get(col, ""), val, op)]
    except ValueError as e:
        return {"status": "error", "reason": str(e)}

    if limit and limit < len(filtered):
        filtered = filtered[:limit]

    return {
        "status": "ok",
        "rows": filtered,
        "count": len(filtered),
        "filtered_count": len(filtered),
        "total_count": total,
    }



@directive("tc-table", "sort", domain_alias="表格", action_aliases={"sort": "排序"})
def tc_table_sort(params: list[str]) -> dict:
    if len(params) < 2:
        return {"status": "error", "reason": "Usage: tc-table;sort,<json_data>,<JSON_query>"}

    rows = _parse_json_rows(params[0])
    if not rows:
        return {"status": "error", "reason": "no data rows provided or JSON parse failed"}
    try:
        query = json.loads(params[1])
    except json.JSONDecodeError:
        return {"status": "error", "reason": f"invalid JSON query: {params[1][:50]}"}

    by = query.get("by")
    if not by:
        return {"status": "error", "reason": "query must have 'by': column name(s)"}

    cols = by if isinstance(by, list) else [by]
    dirs = query.get("dir", "asc")
    if isinstance(dirs, str):
        dirs = [dirs] * len(cols)
    reversed_dirs = [d == "desc" for d in dirs]

    sorted_rows = list(rows)
    for col, rev in zip(reversed(cols), reversed(reversed_dirs)):
        sorted_rows.sort(
            key=lambda r, c=col: (float(r.get(c, 0)) if r.get(c, "").replace("-", "").replace(".", "").isdigit() and r.get(c, "").count(".") <= 1 else r.get(c, "")),
            reverse=rev,
        )

    return {
        "status": "ok",
        "rows": sorted_rows,
        "count": len(sorted_rows),
    }



@directive("tc-table", "pivot", domain_alias="表格", action_aliases={"pivot": "透视"})
def tc_table_pivot(params: list[str]) -> dict:
    if len(params) < 2:
        return {"status": "error", "reason": "Usage: tc-table;pivot,<json_data>,<JSON_query>"}

    rows = _parse_json_rows(params[0])
    if not rows:
        return {"status": "error", "reason": "no data rows provided or JSON parse failed"}
    try:
        query = json.loads(params[1])
    except json.JSONDecodeError:
        return {"status": "error", "reason": f"invalid JSON query: {params[1][:50]}"}

    group_col = query.get("group")
    agg = query.get("agg", "count")
    on_col = query.get("on")

    if not group_col:
        return {"status": "error", "reason": "query must have 'group': column name"}

    if agg != "count" and not on_col:
        return {"status": "error", "reason": f"agg='{agg}' requires 'on': column name"}

    groups = {}
    for r in rows:
        key = r.get(group_col, "")
        if key not in groups:
            groups[key] = []
        groups[key].append(r)

    result_rows = []
    for key, group_rows in groups.items():
        if agg == "count":
            result_rows.append({group_col: key, "count": len(group_rows)})
        else:
            values = []
            for r in group_rows:
                try:
                    values.append(float(r.get(on_col, 0)))
                except (ValueError, TypeError):
                    pass
            if not values:
                continue
            col_name = f"{agg}_{on_col}"
            if agg == "sum":
                val = sum(values)
            elif agg == "avg":
                val = sum(values) / len(values)
            elif agg == "min":
                val = min(values)
            elif agg == "max":
                val = max(values)
            else:
                return {"status": "error", "reason": f"unknown agg: {agg}"}
            result_rows.append({group_col: key, col_name: val})

    return {
        "status": "ok",
        "group": group_col,
        "agg": agg,
        "rows": result_rows,
        "count": len(result_rows),
    }



@directive("tc-table", "join", domain_alias="表格", action_aliases={"join": "合并"})
def tc_table_join(params: list[str]) -> dict:
    if len(params) < 2:
        return {"status": "error", "reason": "Usage: tc-table;join,<json_a>,<json_b>,<JSON_query>"}

    left = _parse_json_rows(params[0])
    right = _parse_json_rows(params[1])

    query = {}
    if len(params) > 2:
        query_str = params[2]
        try:
            query = json.loads(query_str)
        except json.JSONDecodeError:
            return {"status": "error", "reason": f"invalid JSON query: {query_str[:50]}"}

    if not left or not right:
        return {"status": "error", "reason": "could not parse two JSON data sources"}

    join_type = query.get("type", "inner")
    on_col = query.get("on")

    if not on_col:
        return {"status": "error", "reason": "query must have 'on': column name"}

    right_index = {}
    for r in right:
        key = r.get(on_col, "")
        if key not in right_index:
            right_index[key] = []
        right_index[key].append(r)

    left_cols = set()
    if left:
        left_cols = set(left[0].keys())
    right_cols = set()
    if right:
        right_cols = set(right[0].keys())
    conflict_cols = left_cols & right_cols - {on_col}
    right_rename = {c: f"{c}_right" for c in conflict_cols}

    result = []
    matched_left = set()

    for lr in left:
        l_key = lr.get(on_col, "")
        matches = right_index.get(l_key, [])
        if matches:
            for rr in matches:
                joined = dict(lr)
                for k, v in rr.items():
                    out_key = right_rename.get(k, k)
                    if out_key != on_col:
                        joined[out_key] = v
                result.append(joined)
                matched_left.add(l_key)
        elif join_type in ("left", "outer"):
            joined = dict(lr)
            for rc in right_cols:
                out_key = right_rename.get(rc, rc)
                if out_key != on_col:
                    joined[out_key] = ""
            result.append(joined)

    if join_type in ("right", "outer"):
        for rr in right:
            r_key = rr.get(on_col, "")
            if r_key not in matched_left:
                joined = {}
                for lc in left_cols:
                    joined[lc] = ""
                for k, v in rr.items():
                    out_key = right_rename.get(k, k)
                    if out_key != on_col:
                        joined[out_key] = v
                    else:
                        joined[on_col] = v
                result.append(joined)

    return {
        "status": "ok",
        "type": join_type,
        "on": on_col,
        "rows": result,
        "count": len(result),
        "left_count": len(left),
        "right_count": len(right),
    }



@directive("tc-table", "write", domain_alias="表格", action_aliases={"write": "写入"})
def tc_table_write(params: list[str]) -> dict:
    if len(params) < 2:
        return {"status": "error", "reason": "Usage: tc-table;write,<json_data>,<file_path>[,<format>]"}

    rows = _parse_json_rows(params[0])
    if not rows:
        return {"status": "error", "reason": "no data rows provided or JSON parse failed"}
    file_path = params[1]
    if not file_path:
        return {"status": "error", "reason": "no file path provided"}
    fmt = _detect_format(file_path, params[2] if len(params) > 2 else None)
    out_path = _resolve_path(file_path)

    try:
        out_path.parent.mkdir(parents=True, exist_ok=True)

        if not rows:
            col_names = []
        else:
            col_names = list(rows[0].keys())

        if fmt == "xlsx":
            try:
                from openpyxl import Workbook
            except ImportError:
                return {"status": "error", "reason": "openpyxl required for .xlsx. pip install openpyxl"}
            wb = Workbook()
            ws = wb.active
            ws.append(col_names)
            for row in rows:
                ws.append([row.get(c, "") for c in col_names])
            wb.save(str(out_path))
        elif fmt == "tsv":
            with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=col_names, delimiter="\t")
                writer.writeheader()
                writer.writerows(rows)
        else:
            with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=col_names)
                writer.writeheader()
                writer.writerows(rows)

        size = out_path.stat().st_size

        return {
            "status": "ok",
            "path": str(out_path),
            "format": fmt,
            "size": size,
            "rows": len(rows),
        }
    except Exception as e:
        logger.exception("tc-table;write failed: %s", file_path)
        return {"status": "error", "reason": str(e)}
